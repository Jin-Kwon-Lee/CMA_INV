import pandas as pd
from openpyxl import Workbook
import random
from pprint import pprint
import numpy as np

def read_excel_file(file_path, sheet_name):
    try:
        # Excel 파일에서 지정된 시트를 읽어와서 DataFrame으로 변환
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
        return df
    except FileNotFoundError:
        print("파일을 찾을 수 없습니다.")
        return None
    except Exception as e:
        print("오류 발생:", e)
        return None
def _to_excel_df(df, file_path, sheet_name):
    df.to_excel(file_path, sheet_name=sheet_name, index=True)  # index=False로 지정하면 DataFrame의 인덱스는 저장되지 않음


def write_df_to_excel(ws, dataframe, start_row, start_column, save_index=True):
    row_index = start_row  # 행 인덱스 초기화
    
    # 데이터를 저장
    for idx, (index, row) in enumerate(dataframe.iterrows(), start=row_index):
        col_index = start_column
        if save_index:
            ws.cell(row=idx, column=col_index, value=index)
            col_index += 1
        for value in row:
            ws.cell(row=idx, column=col_index, value=value)
            col_index += 1
        row_index += 1


def my_range(start, stop):
    return list(range(start, start + stop))
    
def gen_mandatory_df(df):
    mandatory_col = ['BL NO','MODEL','YR','CHASSISNO.','WEIGHT','CONTAINERNO.','SEAL NO.','TARE','complete']
    df = df[mandatory_col]
    return df

def fill_value (col, value, MODELs):
    if col == 'MODEL': new_value = random.choice(MODELs)
    elif col == 'CHASSISNO.' : new_value = random.randint(100000, 999999)
        
    return new_value

def expand_df(df, num_rows):
    if len(df) < 5:
        new_df = pd.DataFrame({
            'BL NO': df.iloc[0]['BL NO'],
            'MODEL': [np.nan] * (5 - len(df)),
            'YR': [np.nan] * (5 - len(df)),
            'CHASSISNO.': [np.nan] * (5 - len(df)),
            'complete': 'X',
        })
        return pd.concat([df, new_df], ignore_index=True)
    else:
        return df

def fill_loss_df (df, MODELs):
    if len(df) < 5:
        df = expand_df(df,5)

    for index, row in df.iterrows():
        for col in ['MODEL','CHASSISNO.']:
            value = row[col]
            if str(value) == 'nan':
                df.at[index, col] = fill_value(col, value, MODELs)
    return df

    
def _run(df):
    cont_header_df = pd.DataFrame()
    wb = Workbook()
    min_cont_cnt = 5

    BL_list = set(df['BL NO'].dropna())
    BL_list = sorted(BL_list)

    MODEL_list = list(set(df['MODEL'].dropna()))

    for BL in BL_list:
        ws = wb.create_sheet(title= BL)
        cont_total_cnt = 1
        info_cnt = 1
        
        BL_df = df[(df['BL NO'] == BL)]
        BL_sheet_name = BL
        CONT_list = set(BL_df['CONTAINERNO.'].dropna())
        
        for CONT in CONT_list:
            CONT_df = BL_df[(BL_df['CONTAINERNO.'] == CONT)]
            cont_inval_cnt = len(CONT_df[(CONT_df['complete'] == 'X')])
            
            cont_valid_cnt = len(CONT_df) - cont_inval_cnt
            cont_name = CONT
            
            if cont_valid_cnt < min_cont_cnt:
                CONT_df = fill_loss_df(CONT_df, MODEL_list)

    
            sr_num = CONT_df.iloc[0,6]
            tare = CONT_df.iloc[0,7]
            cont_cnt = len(CONT_df)

            
            if (CONT_df['complete'] == 'X').any()  : cont_weight = '7000'
            else : cont_weight = CONT_df['WEIGHT'].astype(int).sum()
            
            cont_header_dict = {
                'CONTAINER': [cont_name],
                'SR_NUM' : [sr_num],
                'COUNT' : [cont_cnt],
                'WEIGHT' : [cont_weight],
                'CBM' : ['50CBM']
            }

            mark_cont_sr_num_dict = {
                'CONTAINER': [cont_name],
                'SR_NUM' : [sr_num]
            }
            mark_tare_dict = {
                'TARE': [tare]
            }

            
            mark_cont_sr_df = pd.DataFrame(mark_cont_sr_num_dict)
            
            cont_header_df = pd.DataFrame(cont_header_dict)
            cont_header_df['COUNT'] = cont_header_df['COUNT'].apply(lambda x: str(x) + 'UNIT')
            cont_header_df['WEIGHT'] = cont_header_df['WEIGHT'].apply(lambda x: str(x) + 'KGS')
           
            mark_tare_df = pd.DataFrame(mark_tare_dict)
            mark_tare_df['TARE'] = mark_tare_df['TARE'].apply(lambda x: str(x) + 'KGS')

            cont_info_df = CONT_df.loc[:,['MODEL','YR','CHASSISNO.'],]
            
            
            cont_info_df.index = my_range(info_cnt, len(cont_info_df))

            write_df_to_excel(ws, mark_cont_sr_df, cont_total_cnt, 1, save_index=False)
            write_df_to_excel(ws, cont_header_df, cont_total_cnt, 5, save_index=False)
            write_df_to_excel(ws, mark_tare_df, cont_total_cnt+1, 1, save_index=False)
            write_df_to_excel(ws, cont_info_df, cont_total_cnt+1, 4, save_index=True)

            cont_total_cnt += (cont_cnt + 1)
            info_cnt += cont_cnt
         
    wb.save('../result_data.xlsx')


excel_file_path = '../input_data.xlsx'  # 읽어올 Excel 파일의 경로
sheet_name = 'raw_data'

df = read_excel_file(excel_file_path,sheet_name)

man_df = gen_mandatory_df(df)
_run(man_df)
